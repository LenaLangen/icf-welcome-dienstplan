import streamlit as st
import pandas as pd
import json
import os
import base64
import requests as req
from datetime import date, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ──────────────────────────────────────────────────────────────────
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
RESPONSES_FILE = os.path.join(DATA_DIR, "responses.json")
TEAM_FILE = os.path.join(DATA_DIR, "team.json")
ADMIN_PIN = "icf2026"

try:
    GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN", "")
    GITHUB_REPO  = st.secrets.get("GITHUB_REPO", "LenaLangen/icf-welcome-dienstplan")
except Exception:
    GITHUB_TOKEN = ""
    GITHUB_REPO  = "LenaLangen/icf-welcome-dienstplan"

# Paare: werden immer zusammen eingeplant
PAIRS = [("Andreas", "Claudia"), ("Jan M.", "Maria M.")]

# Gottesdienstgrössen
SLOT_SIZES = {"09:30": 6, "11:30": 5}

# Einsatz-Grenzen pro Monat
MAX_PER_MONTH = 4
TARGET_PER_MONTH = 3  # Ideal

os.makedirs(DATA_DIR, exist_ok=True)

TL_DEFAULT = ["Anne", "Gitta", "Kerstin", "Lena L.", "Nikita", "Rebecca"]
TEAM_DEFAULT = [
    "Andreas", "Claudia", "David", "Gundula", "Helena", "Hivin",
    "Isabella", "Jan M.", "Jan R.", "Jessie", "Jürgen",
    "Lara", "Laura", "Lena K.", "Maria", "Maria M.",
    "Melanie", "Nathalie", "Nicole", "Ralf", "Sarah", "Sofia",
    "Sophie", "Susanne", "Ute K.", "Ute S.", "Veronika",
]

GERMAN_MONTHS = [
    "", "Januar", "Februar", "März", "April", "Mai", "Juni",
    "Juli", "August", "September", "Oktober", "November", "Dezember"
]

COLORS = {
    "tl_header": "2C5F8A",
    "tl_fill": "D6E8F5",
    "regular_header": "1A6B3A",
    "regular_fill": "D6F0E1",
    "sunday_fill": "F0F0F0",
    "month_header": "1A3C5E",
    "time_fill": "E8E8E8",
}


# ── GitHub storage ────────────────────────────────────────────────────────────

def _gh_headers():
    return {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}

def _gh_read(path):
    """Returns (data_dict, sha) or (None, None)."""
    r = req.get(f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}",
                headers=_gh_headers(), timeout=10)
    if r.status_code == 200:
        j = r.json()
        return json.loads(base64.b64decode(j["content"]).decode()), j["sha"]
    return None, None

def _gh_write(path, data_dict, sha=None):
    """Create or update a file on GitHub. Returns True on success."""
    payload = {
        "message": "update data",
        "content": base64.b64encode(
            json.dumps(data_dict, ensure_ascii=False, indent=2).encode()
        ).decode(),
    }
    if sha:
        payload["sha"] = sha
    r = req.put(f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}",
                headers=_gh_headers(), json=payload, timeout=10)
    return r.status_code in (200, 201)


# ── Data helpers ─────────────────────────────────────────────────────────────

def load_team():
    if GITHUB_TOKEN:
        data, _ = _gh_read("data/team.json")
        if data:
            return data
    if os.path.exists(TEAM_FILE):
        with open(TEAM_FILE) as f:
            return json.load(f)
    return {"tl": TL_DEFAULT, "regular": TEAM_DEFAULT}


def save_team(team):
    if GITHUB_TOKEN:
        _, sha = _gh_read("data/team.json")
        _gh_write("data/team.json", team, sha)
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(TEAM_FILE, "w") as f:
        json.dump(team, f, ensure_ascii=False, indent=2)


def load_responses():
    if GITHUB_TOKEN:
        data, _ = _gh_read("data/responses.json")
        if data:
            return data
    if os.path.exists(RESPONSES_FILE):
        with open(RESPONSES_FILE) as f:
            return json.load(f)
    return {}


def save_responses(responses):
    if GITHUB_TOKEN:
        # Nochmal frisch lesen bevor wir schreiben → verhindert Überschreiben
        # anderer gleichzeitiger Antworten (race condition)
        current, sha = _gh_read("data/responses.json")
        if current:
            # Merge: bestehende Antworten behalten, neue hinzufügen
            for month_key, month_data in responses.items():
                if month_key not in current:
                    current[month_key] = {}
                current[month_key].update(month_data)
            responses = current
        ok = _gh_write("data/responses.json", responses, sha)
        if not ok:
            raise RuntimeError(
                "GitHub-Token ungültig oder abgelaufen. "
                "Bitte Lena kontaktieren — der Token muss in Streamlit Secrets erneuert werden."
            )
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(RESPONSES_FILE, "w") as f:
        json.dump(responses, f, ensure_ascii=False, indent=2)


CONFIG_FILE = os.path.join(DATA_DIR, "config.json")

def load_config():
    if GITHUB_TOKEN:
        data, _ = _gh_read("data/config.json")
        if data:
            return data
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE) as f:
            return json.load(f)
    return {}

def save_config(config):
    if GITHUB_TOKEN:
        _, sha = _gh_read("data/config.json")
        _gh_write("data/config.json", config, sha)
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

def get_active_sundays(year: int, month: int):
    """All Sundays of the month minus manually deactivated ones."""
    config = load_config()
    excluded = config.get(response_key(year, month), {}).get("excluded_sundays", [])
    cal = calendar.monthcalendar(year, month)
    sundays = []
    for week in cal:
        if week[6] != 0:
            d = date(year, month, week[6])
            if d.strftime("%d.%m.") not in excluded:
                sundays.append(d)
    return sundays

def get_sundays(year: int, month: int):
    cal = calendar.monthcalendar(year, month)
    sundays = []
    for week in cal:
        if week[6] != 0:
            sundays.append(date(year, month, week[6]))
    return sundays


def response_key(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


# ── Pages ─────────────────────────────────────────────────────────────────────

SURVEY_DEADLINE_DAY = 25  # Umfrage geschlossen ab dem 26. des Vormonats

def is_survey_open(year: int, month: int) -> tuple[bool, date]:
    """Returns (is_open, deadline_date). Deadline = 25. des Vormonats."""
    first_of_month = date(year, month, 1)
    prev_month_last = first_of_month - timedelta(days=1)
    deadline = date(prev_month_last.year, prev_month_last.month, SURVEY_DEADLINE_DAY)
    return date.today() <= deadline, deadline


def page_survey():
    st.title("🌸 Welcome-Dienst Verfügbarkeiten eintragen")

    team = load_team()
    all_members = sorted(team["tl"] + team["regular"])

    col1, col2 = st.columns(2)
    with col1:
        name = st.selectbox("Dein Name", ["– bitte wählen –"] + all_members)
    with col2:
        today = date.today()
        next_month = today.replace(day=1) + timedelta(days=32)
        years = [today.year, today.year + 1]
        year = st.selectbox("Jahr", years, index=0 if next_month.year == today.year else 1)
        month = st.selectbox("Monat", list(range(1, 13)),
                             format_func=lambda m: GERMAN_MONTHS[m],
                             index=next_month.month - 1)

    open_, deadline = is_survey_open(year, month)
    if not open_:
        st.error(f"⏰ Die Umfrage für {GERMAN_MONTHS[month]} {year} ist geschlossen (Deadline war der {deadline.strftime('%d.%m.%Y')}).")
        return
    else:
        st.info(f"Bitte trage ein, an welchen Sonntagen du eingesetzt werden kannst. **Deadline: {deadline.strftime('%d.%m.%Y')}**")

    if name == "– bitte wählen –":
        st.info("Wähle deinen Namen aus der Liste.")
        return

    is_tl = name in team["tl"]
    sundays = get_active_sundays(year, month)
    config = load_config()
    mkey = response_key(year, month)

    st.markdown(f"### {GERMAN_MONTHS[month]} {year}")

    if is_tl:
        st.info("Du bist Tagesleitung und kannst pro Gottesdienst angeben, ob du als **TL** und/oder als **Welcomer** verfügbar bist.")
    else:
        st.markdown("Wähle für jeden Sonntag aus, wann du Zeit hast:")

    selections = {}
    for sunday in sundays:
        ds = sunday.strftime("%d.%m.")
        st.markdown(f"**{ds}**")
        t930  = get_slot_label(config, mkey, ds, "09:30") + " Uhr"
        t1130 = get_slot_label(config, mkey, ds, "11:30") + " Uhr"

        if is_tl:
            for slot, label in [("09:30", t930), ("11:30", t1130)]:
                cols = st.columns([2, 2, 2, 1])
                cols[0].markdown(f"*{label}*")
                as_tl = cols[1].checkbox("Als Tagesleitung", key=f"{ds}_{slot}_tl")
                as_welcome = cols[2].checkbox("Als Welcomer", key=f"{ds}_{slot}_w")
                if ds not in selections:
                    selections[ds] = {}
                selections[ds][slot] = as_welcome
                selections[ds][f"{slot}_tl"] = as_tl
        else:
            cols = st.columns([2, 2, 2])
            c930  = cols[0].checkbox(t930,  key=f"{ds}_930")
            c1130 = cols[1].checkbox(t1130, key=f"{ds}_1130")
            selections[ds] = {"09:30": c930, "11:30": c1130}

        st.markdown("")

    st.markdown("---")
    note = st.text_area("Anmerkungen (optional)", placeholder="z. B. 'Am 15.06. nur wenn nötig'")

    if st.button("✅ Verfügbarkeit speichern", type="primary"):
        try:
            responses = load_responses()
            key = response_key(year, month)
            if key not in responses:
                responses[key] = {}
            responses[key][name] = {"availability": selections, "note": note}
            save_responses(responses)
            st.success(f"Danke, {name}! Deine Verfügbarkeit für {GERMAN_MONTHS[month]} wurde gespeichert. ✅")
            st.balloons()
        except Exception as e:
            st.error(f"⚠️ Speichern fehlgeschlagen — bitte nochmal versuchen. ({e})")


def page_overview():
    st.title("📋 Übersicht – Rückmeldungen")
    st.caption("Dieser Bereich ist nur für die Koordination. Bitte PIN eingeben.")

    if "admin_ok" not in st.session_state:
        st.session_state.admin_ok = False

    if not st.session_state.admin_ok:
        pin = st.text_input("Admin-PIN", type="password")
        if st.button("Einloggen"):
            if pin == ADMIN_PIN:
                st.session_state.admin_ok = True
                st.rerun()
            else:
                st.error("Falscher PIN.")
        return

    team = load_team()
    all_members = sorted(team["tl"] + team["regular"])
    responses = load_responses()

    today = date.today()
    next_month = today.replace(day=1) + timedelta(days=32)
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Jahr", [today.year, today.year + 1],
                            index=0 if next_month.year == today.year else 1, key="ov_year")
    with col2:
        month = st.selectbox("Monat", list(range(1, 13)),
                             format_func=lambda m: GERMAN_MONTHS[m],
                             index=next_month.month - 1, key="ov_month")

    key = response_key(year, month)
    month_responses = responses.get(key, {})
    sundays = get_active_sundays(year, month)

    responded = set(month_responses.keys())
    not_responded = set(all_members) - responded

    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("✅ Geantwortet", len(responded))
    with col_b:
        st.metric("⏳ Ausstehend", len(not_responded))

    if not_responded:
        with st.expander("Noch keine Rückmeldung von:"):
            for n in sorted(not_responded):
                role = "TL" if n in team["tl"] else "Welcome"
                st.write(f"- {n} ({role})")

    if not month_responses:
        st.info("Noch keine Rückmeldungen für diesen Monat.")
        return

    st.markdown(f"### Verfügbarkeitsmatrix – {GERMAN_MONTHS[month]} {year}")
    st.caption("TL = Tagesleitung · W = Welcomer · TL+W = beides · — = nicht verfügbar")

    config = load_config()
    mkey = response_key(year, month)
    rows = []
    for n in sorted(month_responses.keys()):
        row = {"Name": n, "Rolle": "TL" if n in team["tl"] else "Welcome"}
        avail = month_responses[n].get("availability", {})
        for sunday in sundays:
            ds = sunday.strftime("%d.%m.")
            for slot in ["09:30", "11:30"]:
                as_tl = avail.get(ds, {}).get(f"{slot}_tl", False)
                as_w  = avail.get(ds, {}).get(slot, False)
                if as_tl and as_w:
                    val = "TL+W"
                elif as_tl:
                    val = "TL"
                elif as_w:
                    val = "W"
                else:
                    val = "—"
                label = get_slot_label(config, mkey, ds, slot)
                row[f"{ds} {label}"] = val
        row["Anmerkung"] = month_responses[n].get("note", "")
        rows.append(row)

    df = pd.DataFrame(rows)
    st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown("---")
    if st.button("📊 Dienstplan erstellen & Excel exportieren", type="primary"):
        schedule = build_schedule(year, month, month_responses, team, sundays)
        path = export_excel(year, month, schedule, sundays)
        with open(path, "rb") as f:
            st.download_button(
                "⬇️ Dienstplan herunterladen",
                data=f,
                file_name=os.path.basename(path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        st.success("Dienstplan erstellt!")
        display_schedule(schedule, sundays, config, mkey)


def page_team():
    st.title("👥 Team verwalten")

    if "admin_ok" not in st.session_state or not st.session_state.admin_ok:
        st.warning("Bitte zuerst im Übersicht-Tab einloggen.")
        return

    team = load_team()

    st.markdown("### Tagesleitung (TL)")
    st.caption("TL-Personen können sich in der Umfrage auch als Welcomer eintragen.")
    tl_str = st.text_area("TL-Personen (eine pro Zeile)",
                           value="\n".join(team["tl"]), height=180)

    st.markdown("### Welcome-Team (einfacher Dienst)")
    reg_str = st.text_area("Team-Mitglieder (eine pro Zeile)",
                           value="\n".join(team["regular"]), height=300)

    if st.button("💾 Speichern", type="primary"):
        team["tl"] = [n.strip() for n in tl_str.splitlines() if n.strip()]
        team["regular"] = [n.strip() for n in reg_str.splitlines() if n.strip()]
        save_team(team)
        st.success("Team gespeichert!")


# ── Schedule builder ──────────────────────────────────────────────────────────

def build_schedule(year, month, responses, team, sundays):
    """
    Regeln:
    - TL wird PRO SLOT zugeteilt (09:30 und 11:30 können verschiedene TLs haben)
    - Bevorzugt: selbe TL für beide Slots am gleichen Sonntag
    - Jede TL mind. 1× pro Monat (falls verfügbar)
    - 09:30: 1 TL + 6 Welcomer | 11:30: 1 TL + 5 Welcomer
    - Personen können in beiden Slots desselben Sonntags stehen
    - Max 4 Einsätze/Person/Monat, Ziel 2–3
    - Paare (Andreas+Claudia, Jan M.+Maria M.) gemeinsam einteilen wenn möglich
    """
    schedule = {}
    counts = {n: 0 for n in team["tl"] + team["regular"]}
    tl_count = {n: 0 for n in team["tl"]}

    def avail_for(name, ds, slot, as_tl=False):
        key = f"{slot}_tl" if as_tl else slot
        return responses.get(name, {}).get("availability", {}).get(ds, {}).get(key, False)

    sunday_ds = [s.strftime("%d.%m.") for s in sundays]

    # ── Schritt 1: TL pro Slot zuteilen ──────────────────────────────────────
    # (ds, slot) → tl_name
    tl_slot = {}

    # Sonntage nach Engpass sortieren (wenigste TL-Verfügbarkeit zuerst)
    def tl_avail_count(ds):
        return sum(1 for n in team["tl"]
                   if avail_for(n, ds, "09:30", True) or avail_for(n, ds, "11:30", True))

    for ds in sorted(sunday_ds, key=tl_avail_count):
        avail_both  = [n for n in team["tl"]
                       if avail_for(n, ds, "09:30", True) and avail_for(n, ds, "11:30", True)]
        avail_930   = [n for n in team["tl"] if avail_for(n, ds, "09:30", True)]
        avail_1130  = [n for n in team["tl"] if avail_for(n, ds, "11:30", True)]

        # Bevorzuge unbesetzte TLs
        def pick(candidates):
            unassigned = [n for n in candidates if tl_count[n] == 0]
            pool = unassigned or candidates
            return min(pool, key=lambda n: tl_count[n]) if pool else None

        if avail_both:
            # Selbe TL für beide Slots → bevorzugt
            chosen = pick(avail_both)
            tl_slot[(ds, "09:30")] = chosen
            tl_slot[(ds, "11:30")] = chosen
            tl_count[chosen] += 1
            counts[chosen] += 1
        else:
            # Separate TLs pro Slot
            c930 = pick(avail_930)
            if c930:
                tl_slot[(ds, "09:30")] = c930
                tl_count[c930] += 1
                counts[c930] += 1
            c1130 = pick([n for n in avail_1130 if n != c930])
            if not c1130:
                c1130 = pick(avail_1130)
            if c1130:
                tl_slot[(ds, "11:30")] = c1130
                if c1130 != c930:
                    tl_count[c1130] += 1
                    counts[c1130] += 1

    # ── Schritt 2: Welcomer pro Slot ────────────────────────────────────────
    for sunday in sundays:
        ds = sunday.strftime("%d.%m.")
        schedule[ds] = {}

        for slot in ["09:30", "11:30"]:
            tl_name = tl_slot.get((ds, slot), "– fehlt –")
            target  = SLOT_SIZES[slot]

            # Pool: verfügbar, nicht die TL dieses Slots, unter Max-Grenze
            pool = [
                n for n in team["regular"] + team["tl"]
                if n != tl_name
                and avail_for(n, ds, slot)
                and counts.get(n, 0) < MAX_PER_MONTH
            ]

            chosen = []
            used   = set()

            # Paare gemeinsam einteilen wenn beide verfügbar
            for p1, p2 in PAIRS:
                if p1 in pool and p2 in pool and len(chosen) + 2 <= target:
                    chosen += [p1, p2]
                    used   |= {p1, p2}

            # Restliche Plätze: wer unter Ziel-Zahl liegt hat Vorrang
            individuals = sorted(
                [n for n in pool if n not in used],
                key=lambda n: (1 if counts.get(n, 0) >= TARGET_PER_MONTH else 0,
                               counts.get(n, 0))
            )
            chosen += individuals[: target - len(chosen)]

            # Doppeleinträge verhindern
            seen   = set()
            chosen = [n for n in chosen if not (n in seen or seen.add(n))]

            for n in chosen:
                counts[n] = counts.get(n, 0) + 1

            schedule[ds][slot] = {"tl": tl_name, "team": chosen}

    return schedule


def display_schedule(schedule, sundays, config, mkey):
    st.markdown("### Vorschau Dienstplan")
    notes = config.get(mkey, {}).get("notes", {})
    for sunday in sundays:
        ds = sunday.strftime("%d.%m.")
        note = notes.get(ds, "")
        header = f"**{ds}**" + (f"  —  *{note}*" if note else "")
        st.markdown(header)
        for slot in ["09:30", "11:30"]:
            slot_data = schedule.get(ds, {}).get(slot, {})
            tl = slot_data.get("tl", "—")
            members = ", ".join(slot_data.get("team", []))
            col1, col2 = st.columns([1, 3])
            label = get_slot_label(config, mkey, ds, slot)
            col1.markdown(f"*{label}*")
            col2.markdown(f"TL: **{tl}** | Team: {members or '—'}")


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(year, month, schedule, sundays):
    config = load_config()
    mkey = response_key(year, month)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{GERMAN_MONTHS[month]} {year}"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value="", bold=False, bg=None, fg="000000", center=False, size=10):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", bold=bold, color=fg, size=size)
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        if center:
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        return c

    # Month header spanning all columns
    total_cols = 1 + len(sundays) * 2
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=f"Welcome-Dienstplan – {GERMAN_MONTHS[month]} {year}")
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    c.fill = PatternFill("solid", start_color=COLORS["month_header"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Date headers + time slot sub-headers
    col = 2
    for sunday in sundays:
        ds = sunday.strftime("%d.%m.")
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        cell(2, col, ds, bold=True, bg=COLORS["sunday_fill"], center=True)
        t930  = get_slot_label(config, mkey, ds, "09:30") + " Uhr"
        t1130 = get_slot_label(config, mkey, ds, "11:30") + " Uhr"
        cell(3, col,     t930,  bold=True, bg=COLORS["time_fill"], center=True)
        cell(3, col + 1, t1130, bold=True, bg=COLORS["time_fill"], center=True)
        col += 2

    cell(2, 1, "", bg=COLORS["sunday_fill"])
    cell(3, 1, "Rolle", bold=True, bg=COLORS["time_fill"], center=True)
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # TL row
    row = 4
    cell(row, 1, "Tagesleitung", bold=True, bg=COLORS["tl_fill"],
         fg=COLORS["tl_header"], center=True)
    ws.row_dimensions[row].height = 20
    col = 2
    for sunday in sundays:
        ds = sunday.strftime("%d.%m.")
        for slot in ["09:30", "11:30"]:
            name = schedule.get(ds, {}).get(slot, {}).get("tl", "—")
            cell(row, col, name, bg=COLORS["tl_fill"], center=True)
            col += 1

    # Team member rows (max of 09:30 and 11:30 slot sizes)
    max_rows = max(SLOT_SIZES.values())
    for i in range(1, max_rows + 1):
        row = 4 + i
        cell(row, 1, str(i), bold=True, bg=COLORS["regular_fill"],
             fg=COLORS["regular_header"], center=True)
        ws.row_dimensions[row].height = 18
        col = 2
        for sunday in sundays:
            ds = sunday.strftime("%d.%m.")
            for slot in ["09:30", "11:30"]:
                members = schedule.get(ds, {}).get(slot, {}).get("team", [])
                name = members[i - 1] if i - 1 < len(members) else ""
                cell(row, col, name, center=True)
                col += 1

    # Notes row (Abendmahl, Taufe, etc.)
    notes = config.get(mkey, {}).get("notes", {})
    if any(notes.get(s.strftime("%d.%m.")) for s in sundays):
        note_row = 4 + max_rows + 1
        ws.row_dimensions[note_row].height = 18
        cell(note_row, 1, "", bg="FFFFFF")
        col = 2
        for sunday in sundays:
            ds = sunday.strftime("%d.%m.")
            note = notes.get(ds, "")
            # Merge over both slots
            ws.merge_cells(start_row=note_row, start_column=col,
                           end_row=note_row, end_column=col + 1)
            c = ws.cell(row=note_row, column=col, value=note)
            c.font = Font(name="Arial", italic=True, color="666666", size=9)
            c.alignment = Alignment(horizontal="center", vertical="center")
            col += 2

    ws.column_dimensions["A"].width = 16
    for i in range(2, 2 + len(sundays) * 2):
        ws.column_dimensions[get_column_letter(i)].width = 14

    path = os.path.join(DATA_DIR, f"Dienstplan_{GERMAN_MONTHS[month]}_{year}.xlsx")
    wb.save(path)
    return path


def get_slot_label(config, month_key, ds, slot):
    """Returns the display time for a slot, custom or default."""
    return config.get(month_key, {}).get("custom_times", {}).get(ds, {}).get(slot, slot)


def page_settings():
    st.title("⚙️ Monats-Einstellungen")
    st.markdown("Sonntage deaktivieren oder Uhrzeiten anpassen — Änderungen gelten sofort für Umfrage und Dienstplan.")

    if "admin_ok" not in st.session_state or not st.session_state.admin_ok:
        st.warning("Bitte zuerst im Übersicht-Tab einloggen.")
        return

    today = date.today()
    next_month = today.replace(day=1) + timedelta(days=32)
    col1, col2 = st.columns(2)
    with col1:
        year = st.selectbox("Jahr", [today.year, today.year + 1],
                            index=0 if next_month.year == today.year else 1, key="cfg_year")
    with col2:
        month = st.selectbox("Monat", list(range(1, 13)),
                             format_func=lambda m: GERMAN_MONTHS[m],
                             index=next_month.month - 1, key="cfg_month")

    all_sundays = get_sundays(year, month)
    config = load_config()
    mkey = response_key(year, month)
    excluded = config.get(mkey, {}).get("excluded_sundays", [])
    custom_times = config.get(mkey, {}).get("custom_times", {})

    st.markdown(f"### Sonntage im {GERMAN_MONTHS[month]} {year}")

    new_excluded = []
    new_custom_times = {}
    new_notes = {}

    for sunday in all_sundays:
        ds = sunday.strftime("%d.%m.")
        active = st.checkbox(f"**{ds}** – Gottesdienst findet statt",
                             value=(ds not in excluded), key=f"cfg_active_{ds}")
        if not active:
            new_excluded.append(ds)
        else:
            cols = st.columns([1, 2, 2, 3])
            cols[0].markdown("Uhrzeiten:")
            t930  = cols[1].text_input("1. GD", key=f"cfg_t930_{ds}",
                                        value=custom_times.get(ds, {}).get("09:30", "09:30"))
            t1130 = cols[2].text_input("2. GD", key=f"cfg_t1130_{ds}",
                                        value=custom_times.get(ds, {}).get("11:30", "11:30"))
            note  = cols[3].text_input("Bemerkung (z. B. Abendmahl, Taufe)",
                                        key=f"cfg_note_{ds}",
                                        value=config.get(mkey, {}).get("notes", {}).get(ds, ""))
            if t930 != "09:30" or t1130 != "11:30":
                new_custom_times[ds] = {"09:30": t930, "11:30": t1130}
            if note:
                new_notes[ds] = note
        st.markdown("")

    if st.button("💾 Speichern", type="primary"):
        if mkey not in config:
            config[mkey] = {}
        config[mkey]["excluded_sundays"] = new_excluded
        config[mkey]["custom_times"] = new_custom_times
        config[mkey]["notes"] = new_notes
        save_config(config)
        st.success("Gespeichert! ✅")


# ── Main ──────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="ICF Welcome – Dienstplan",
    page_icon="🙌",
    layout="wide",
)

st.sidebar.title("🫶 ICF Welcome-Team")
st.sidebar.markdown("---")

page = st.sidebar.radio(
    "Navigation",
    ["🌸 Verfügbarkeiten eintragen", "📋 Übersicht & Dienstplan", "👥 Team verwalten", "⚙️ Monats-Einstellungen"],
)

if page == "🌸 Verfügbarkeiten eintragen":
    page_survey()
elif page == "📋 Übersicht & Dienstplan":
    page_overview()
elif page == "👥 Team verwalten":
    page_team()
elif page == "⚙️ Monats-Einstellungen":
    page_settings()

st.sidebar.markdown("---")
st.sidebar.caption("ICF Welcome-Team Tool · v1.2")
